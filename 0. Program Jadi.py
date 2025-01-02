import tkinter as tk
from tkinter import messagebox
import qrcode
from openpyxl import Workbook, load_workbook
from datetime import datetime
import cv2
from pyzbar.pyzbar import decode
import os

# Fungsi untuk login
def login():
    username = entry_username.get()
    password = entry_password.get()

    if username == "1" and password == "1":
        open_main_menu()
    else:
        messagebox.showerror("Login Error", "Username atau password salah")

# Fungsi untuk membuka menu utama
def open_main_menu():
    login_frame.pack_forget()
    main_menu_frame.pack()

# Fungsi untuk membuat kode QR dan menyimpan sebagai gambar dan teks dalam Excel
def buat_kode_qr_dan_simpan(text_input):
    if not os.path.exists("qr"):
        os.makedirs("qr")
    if not os.path.exists("excel"):
        os.makedirs("excel")
    
    qr = qrcode.make(text_input)
    words = text_input.split(" | ")
    nama_barang = words[0]
    qr_filename = f"qr/{nama_barang}.png"
    qr.save(qr_filename)

    excel_filename = "excel/Data_Barang_Toko.xlsx"
    if os.path.exists(excel_filename):
        workbook = load_workbook(excel_filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DATA BARANG"
        sheet['A1'] = "Nama Barang"
        sheet['B1'] = "Harga Barang"
    
    row = sheet.max_row + 1
    words = text_input.split(" | ")
    for col, word in enumerate(words, start=1):
        sheet.cell(row=row, column=col, value=word)
    
    workbook.save(excel_filename)

def tambah_barang():
    main_menu_frame.pack_forget()
    form_tambah_barang_frame.pack()

def simpan_barang():
    nama_barang = entry_nama_barang.get()
    harga_barang = entry_harga_barang.get()

    if nama_barang and harga_barang:
        text_input = f"{nama_barang} | {harga_barang}"
        buat_kode_qr_dan_simpan(text_input)
        messagebox.showinfo("Tambah Barang", "Barang berhasil ditambahkan dan QR code telah dibuat.")
        form_tambah_barang_frame.pack_forget()
        main_menu_frame.pack()
    else:
        messagebox.showerror("Input Error", "Masukkan nama dan harga barang terlebih dahulu.")

# Fungsi Euclidean untuk menghitung GCD
def extended_gcd(a, b):
    if a == 0:
        return b, 0, 1
    gcd_val, x1, y1 = extended_gcd(b % a, a)
    x = y1 - (b // a) * x1
    y = x1
    return gcd_val, x, y

# Mencari nilai d untuk dekripsi
def calculate_d(e, qn):
    gcd_val, d, _ = extended_gcd(e, qn)
    if gcd_val != 1:
        raise ValueError("e dan qn tidak relatif prima. Tidak ada invers.")
    return d % qn

# Fungsi untuk enkripsi data hasil scan
def encrypt_message(message, e, n):
    encrypted_message = ""
    for char in message:
        if char.isalpha():
            m = ord(char.lower()) - ord('a') + 7
        elif char.isdigit():
            m = int(char) + 33
        else:
            m = 0
        c = pow(m, e, n)
        encrypted_message += f"{c:04d}"
    return encrypted_message

# Fungsi untuk menyimpan data hasil scan ke Excel dengan enkripsi
def simpan_ke_excel(data, jumlah_barang, e, n):
    if not os.path.exists("excel"):
        os.makedirs("excel")

    excel_filename = "excel/Encrypted_Data.xlsx"
    if os.path.exists(excel_filename):
        workbook = load_workbook(excel_filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data Penjualan"
        sheet.append(["Data Enkripsi", "Tanggal", "Waktu"])

    parts = data.split('|')
    if len(parts) >= 2:
        barang = parts[0].strip()
        try:
            harga = int(parts[1].strip())
        except ValueError:
            print("Format harga tidak valid.")
            return
    else:
        print("Data QR tidak lengkap.")
        return

    total_harga = harga * jumlah_barang
    tanggal = datetime.now().strftime('%Y-%m-%d')
    waktu = datetime.now().strftime('%H:%M:%S')
    message = f"{barang} {harga} {total_harga} {jumlah_barang} {tanggal} {waktu}"
    data_terenkripsi = encrypt_message(message, e, n)
    sheet.append([data_terenkripsi, tanggal, waktu])
    workbook.save(excel_filename)

# Fungsi untuk membuka pemindaian QR Code
def scan_qr_code():
    try:
        jumlah_barang = int(entry_jumlah_barang.get())
    except ValueError:
        messagebox.showerror("Input Error", "Masukkan jumlah barang dalam angka.")
        return

    p, q, e = 71, 79, 89
    n = p * q
    qn = (p - 1) * (q - 1)
    d = calculate_d(e, qn)

    cap = cv2.VideoCapture(0)
    data_terdeteksi = set()
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        for barcode in decode(frame):
            data = barcode.data.decode('utf-8')
            if data not in data_terdeteksi:
                start_time = datetime.now()  # Mulai waktu enkripsi
                simpan_ke_excel(data, jumlah_barang, e, n)
                end_time = datetime.now()  # Akhiri waktu enkripsi
                elapsed_time = (end_time - start_time).microseconds / 1000  # Hitung waktu dalam ms
                label_kecepatan_enkripsi.config(text=f"Kecepatan Enkripsi: {elapsed_time:.2f} ms")  # Update label
                data_terdeteksi.add(data)
                cap.release()
                cv2.destroyAllWindows()
                messagebox.showinfo("Scan Success", "Data berhasil disimpan.")
                return
        cv2.imshow("QR Code Scanner", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()

# Fungsi Euclidean untuk menghitung FPB pada decrypt
def fpb(angka1, angka2):
    while angka2 != 0:
        angka1, angka2 = angka2, angka1 % angka2
    return angka1

def fpb_ekstended(angka1, angka2):
    if angka1 == 0:
        return angka2, 0, 1
    nilai_fpb, x1, y1 = fpb_ekstended(angka2 % angka1, angka1)
    xx = y1 - (angka2 // angka1) * x1
    yy = x1
    return nilai_fpb, xx, yy

def hitung_dd(ee, qn):
    nilai_fpb, dd, _ = fpb_ekstended(ee, qn)
    if nilai_fpb != 1:
        raise ValueError("ee dan qn tidak relatif prima. Tidak ada invers.")
    return dd % qn

# Fungsi untuk dekripsi pesan terenkripsi
def dekripsi_pesan(pesan_terenkripsi, dd, nn):
    pesan_dekripsi = ""
    for i in range(0, len(pesan_terenkripsi), 4):
        cc = int(pesan_terenkripsi[i:i+4])
        mm = pow(cc, dd, nn)
        if 7 <= mm <= 32:
            pesan_dekripsi += chr(mm + ord('a') - 7)
        elif 33 <= mm <= 42:
            pesan_dekripsi += chr(mm - 33 + ord('0'))
        else:
            pesan_dekripsi += ' '
    return pesan_dekripsi

def simpan_hasil_dekripsi(pesan_dekripsi):
    bagian = pesan_dekripsi.split()
    if len(bagian) < 8:
        messagebox.showerror("Kesalahan", "Data hasil dekripsi tidak lengkap.")
        return

    nama_barang = bagian[0]
    harga = bagian[1]
    total_harga = bagian[2]
    jumlah = bagian[3]
    tanggal = f"{bagian[4]}-{bagian[5]}-{bagian[6]}"
    waktu = f"{bagian[7]}:{bagian[8]}:{bagian[9]}"

    if not os.path.exists("excel"):
        os.makedirs("excel")
    nama_file_excel = "excel/Data_Penjualan.xlsx"
    if os.path.exists(nama_file_excel):
        buku_kerja = load_workbook(nama_file_excel)
        lembar = buku_kerja.active
    else:
        buku_kerja = Workbook()
        lembar = buku_kerja.active
        lembar.title = "Data Penjualan"
        lembar.append(["Nama Barang", "Harga", "Total Harga", "Jumlah", "Tanggal", "Waktu"])

    lembar.append([nama_barang, harga, total_harga, jumlah, tanggal, waktu])
    buku_kerja.save(nama_file_excel)
    messagebox.showinfo("Sukses", "Hasil dekripsi berhasil disimpan.")

# Fungsi untuk membuka frame dekripsi
def buka_frame_dekripsi():
    main_menu_frame.pack_forget()
    frame_dekripsi.pack()

# Fungsi input dekripsi
def proses_dekripsi():
    try:
        pp = int(masukkan_pp.get())
        qq = int(masukkan_qq.get())
        ee = 89
        pesan_terenkripsi = masukkan_enkripsi.get()
    except ValueError:
        messagebox.showerror("Kesalahan Input", "Pastikan semua input valid.")
        return
    
    try:
        nn = pp * qq
        qn = (pp - 1) * (qq - 1)
        dd = hitung_dd(ee, qn)
    except ValueError as err:
        messagebox.showerror("Kesalahan Kunci", str(err))
        return

    start_time = datetime.now()  # Mulai waktu dekripsi
    pesan_dekripsi = dekripsi_pesan(pesan_terenkripsi, dd, nn)
    end_time = datetime.now()  # Akhiri waktu dekripsi
    elapsed_time = (end_time - start_time).microseconds / 1000  # Hitung waktu dalam ms
    label_kecepatan_dekripsi.config(text=f"Kecepatan Dekripsi: {elapsed_time:.2f} ms")  # Update label

    simpan_hasil_dekripsi(pesan_dekripsi)
    frame_dekripsi.pack_forget()
    main_menu_frame.pack()

def open_scan_barang():
    main_menu_frame.pack_forget()
    scan_frame.pack()

# Fungsi untuk kembali ke menu utama
def kembali_ke_menu_utama():
    form_tambah_barang_frame.pack_forget()
    scan_frame.pack_forget()
    frame_dekripsi.pack_forget()
    main_menu_frame.pack()


# GUI
root = tk.Tk()
root.title("Aplikasi Enkripsi Dekripsi Toko retail")
root.geometry("400x300")

login_frame = tk.Frame(root)
login_frame.pack(pady=50)
label_username = tk.Label(login_frame, text="Username:")
label_username.grid(row=0, column=0, pady=5)
entry_username = tk.Entry(login_frame)
entry_username.grid(row=0, column=1, pady=5)

label_password = tk.Label(login_frame, text="Password:")
label_password.grid(row=1, column=0, pady=5)
entry_password = tk.Entry(login_frame, show="*")
entry_password.grid(row=1, column=1, pady=5)

button_login = tk.Button(login_frame, text="Login", command=login)
button_login.grid(row=2, columnspan=2, pady=10)

main_menu_frame = tk.Frame(root)
button_tambah_barang = tk.Button(main_menu_frame, text="Tambah Barang", command=tambah_barang)
button_tambah_barang.pack(pady=10)

button_scan_barang = tk.Button(main_menu_frame, text="Scan Barang", command=open_scan_barang)
button_scan_barang.pack(pady=10)

tombol_dekripsi = tk.Button(main_menu_frame, text="Dekripsi", command=buka_frame_dekripsi)
tombol_dekripsi.pack(pady=10)

form_tambah_barang_frame = tk.Frame(root)
label_nama_barang = tk.Label(form_tambah_barang_frame, text="Nama Barang:")
label_nama_barang.grid(row=0, column=0, pady=5)
entry_nama_barang = tk.Entry(form_tambah_barang_frame)
entry_nama_barang.grid(row=0, column=1, pady=5)

label_harga_barang = tk.Label(form_tambah_barang_frame, text="Harga Barang:")
label_harga_barang.grid(row=1, column=0, pady=5)
entry_harga_barang = tk.Entry(form_tambah_barang_frame)
entry_harga_barang.grid(row=1, column=1, pady=5)

button_simpan_barang = tk.Button(form_tambah_barang_frame, text="Simpan Barang", command=simpan_barang)
button_simpan_barang.grid(row=2, columnspan=2, pady=10)

button_kembali_tambah_barang = tk.Button(form_tambah_barang_frame, text="Kembali", command=kembali_ke_menu_utama)
button_kembali_tambah_barang.grid(row=3, columnspan=2, pady=10)

scan_frame = tk.Frame(root)
label_jumlah_barang = tk.Label(scan_frame, text="Masukkan Jumlah Barang:")
label_jumlah_barang.grid(row=0, column=0, pady=5)
entry_jumlah_barang = tk.Entry(scan_frame)
entry_jumlah_barang.grid(row=0, column=1, pady=5)

button_mulai_scan = tk.Button(scan_frame, text="Scan Barang", command=scan_qr_code)
button_mulai_scan.grid(row=1, columnspan=2, pady=10)

button_kembali_scan = tk.Button(scan_frame, text="Kembali", command=kembali_ke_menu_utama)
button_kembali_scan.grid(row=2, columnspan=2, pady=10)

frame_dekripsi = tk.Frame(root)
label_enkripsi = tk.Label(frame_dekripsi, text="Enkripsi:")
label_enkripsi.grid(row=0, column=0, pady=5)
masukkan_enkripsi = tk.Entry(frame_dekripsi, width=30)
masukkan_enkripsi.grid(row=0, column=1, pady=5)

label_pp = tk.Label(frame_dekripsi, text="Masukkan kunci pertama:")
label_pp.grid(row=1, column=0, pady=5)
masukkan_pp = tk.Entry(frame_dekripsi)
masukkan_pp.grid(row=1, column=1, pady=5)

label_qq = tk.Label(frame_dekripsi, text="Masukkan kunci kedua:")
label_qq.grid(row=2, column=0, pady=5)
masukkan_qq = tk.Entry(frame_dekripsi)
masukkan_qq.grid(row=2, column=1, pady=5)

tombol_proses_dekripsi = tk.Button(frame_dekripsi, text="Dekripsi", command=proses_dekripsi)
tombol_proses_dekripsi.grid(row=4, columnspan=2, pady=10)

button_kembali_dekripsi = tk.Button(frame_dekripsi, text="Kembali", command=kembali_ke_menu_utama)
button_kembali_dekripsi.grid(row=5, columnspan=2, pady=10)

# Tambahkan variabel untuk indikator kecepatan
label_kecepatan_enkripsi = tk.Label(scan_frame, text="Kecepatan Enkripsi: 0 ms")
label_kecepatan_enkripsi.grid(row=3, columnspan=2, pady=5)

# Tambahkan variabel untuk indikator kecepatan dekripsi
label_kecepatan_dekripsi = tk.Label(frame_dekripsi, text="Kecepatan Dekripsi: 0 ms")
label_kecepatan_dekripsi.grid(row=6, columnspan=2, pady=5)

root.mainloop()

